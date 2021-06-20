# Declaring Color Values
class Color:
    PURPLE = '\033[95m'
    CYAN = '\033[96m'
    DARKCYAN = '\033[36m'
    BLUE = '\033[94m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
    END = '\033[0m'


# Importing Libraries
try:
    from tabulate import tabulate
    import datetime as datetime
    from tqdm import tqdm as progressbar  # Visual using progress bar
    import time
    import pyfiglet as art
    import openpyxl as excel
    from sys import exit
    from twilio.rest import Client
except:
    print(Color.RED + Color.BOLD + 'ERROR:\nPython Packages not found!' + Color.END)
    print("\nInstall Python Packages to get started!")
    exit()

name = str(input("\nEnter Your Name:"))
name = name.capitalize()
print('\nHi,', Color.BOLD + name.capitalize() + Color.END)

# SMS API Integration with Twilio Trial
sendto = "+91" + input('\nEnter your phone number:')
sid = 'ACc514b4800253a9b8c56867d650b2e8d2'
auth_token = 'a124c1390016e1e505856a7b66b02889'
client = Client(sid, auth_token)

passcode = 0x75AB  # Password in HEX ('0x' - tells python value is in hex)

count = 0  # Count user's invalid attempts
user_pass = ""  # Value received by user

# Password Check
while user_pass != passcode and count < 3:  # Ask password from user thrice
    try:
        user_pass = int(input('Enter PIN:\n'))
    except:
        print(
            "\nERROR\n•Password can't consists of alphabets.\n•Password must be 5 digits.\n")  # If the value isn't numeric

    if user_pass == passcode:
        print('Access granted for', name.capitalize() + ".\n")
        break  # Let's the loop end & starts new block
    else:
        print('Access denied. Try again,' + name.capitalize() + ".")
        count += 1
        print('You have', 3 - count, 'counts left.\n')
        if count == 3:
            print('Let the brain rest,', name.lower(), ':)')
            print(Color.BOLD + 'USER LOCKED OUT!!' + Color.END)
            resp = client.messages.create(body="\nSuspicious User Found in Amul Billing!\nName:" + name.capitalize()+'\nPhone:' + sendto, from_='+13126754624',to='+919931004934')
            exit()
            # Throws user out of script

print(Color.BOLD + 'WELCOME TO BILLING MACHINE, ' + name.upper() + Color.END)

try:
    resp = client.messages.create(body="Welcome to Sahil's program, " + name.capitalize(), from_='+13126754624',
                                  to=sendto)
except:
    print('Phone Number invalid or not verified!')

rate = [['Gold', '500ml', 26.18],
        ['Gold', '1L', 51.35],
        ['Shakti', '500ml', 23.20],
        ['Shakti', '1L', 45.40],
        ['Cow', '500ml', 22.20],
        ['Cow', '1L', 43.40],
        ['Taaza', '500ml', 21.18],
        ['Taaza', '1L', 41.40],
        ['Amul Kool (Plastic)', '200ml * 30', 540.00],
        ['Dahi', '200g', 13.86],
        ['Dahi', '400g', 24.63],
        ['Dahi', '1KG', 56.80],
        ['Lassi (Packet)', '180ml', 9.10],
        ['Paneer', '200g', 66.00],
        ['Paneer', '1KG', 315.00]
        ]  # Rate of products

ans1 = input('Want to know rates? [Y/N] \n')
if (ans1.upper() == "Y") or (ans1 == "1"):
    print(tabulate(rate, headers=[Color.BOLD + "Item", "Quantity", "Selling Price" + Color.END], numalign="left"))
    print('\n#Exclusively_On_Softline')
else:
    print('Okay,', name, ':)')

ans2 = input('\nWant to start billing machine? [Y/N]\n')

if (ans2.upper() == "Y") or (ans2 == "1"):

    # Progress Bar for vibes...
    for i in progressbar(range(10),
                         desc='Opening...',
                         ascii=False, ncols=75):
        time.sleep(0.25)
    print("Complete.\n")

    # Let's create art
    banner = art.figlet_format("WELCOME TO SOFTLINE AMUL BILLING", font="digital")
    print(Color.BOLD + Color.GREEN + banner + Color.END)

    # Excel Connection Starts Here!
    print('Connect to EXCEL of date:\n')
    print(Color.BOLD + "[1] Today\n"
                       "[2] Yesterday\n"
                       "[3] Enter Date Manually\n"
                       "[4] USE SAMPLE FILE\n" + Color.END)
    ans3 = int(input("Enter 1 or 2 or 3:\n"))
    dat = ''
    if ans3 == 1:
        dat = today = datetime.date.today().strftime("%d-%m-%Y")
    if ans3 == 2:
        yesterday = datetime.date.today() - datetime.timedelta(days=1)
        dat = yesterday.strftime("%d-%m-%Y")
    if ans3 == 3:
        dat = manual_date = str(input("\nEnter Date Manually (DD-MM-YYYY):"))
    if ans3 == 4:
        dat = "15-06-2021"
        print("Sample Excel File Date is: " + dat)
        print("Connecting...\n")
    try:
        excel_path = 'O-' + dat + '.xlsx'
        wb = excel.load_workbook("Excel Files/" + excel_path)
    except:
        print(Color.RED + "ERROR\n"
                          "Can't find Excel sheet with date - " + dat + ".\n"
                                                                        "Please Enter Excel File Location Manually." + Color.END)
        excel_path = input("\n\nEnter Excel File Path \n(without quotes)(with double slashes):\n")
        wb = excel.load_workbook(excel_path)
    try:
        print('Connected with Excel - ' + Color.BOLD + excel_path + Color.END)
        print('Connecting with sheets inside Excel...')
        sh1 = wb["A1 KTRA-IND"]
        sh2 = wb["A2 GANJ"]
        sh3 = wb["A3 BIDUPUR"]
        sh4 = wb['A4 MARAI']
        print(Color.BLUE + Color.BOLD + 'File successfully connected.\n' + Color.END)
    except:
        print(Color.RED + "ERROR:\n File is invalid!" + Color.END)

    # After Excel Connected
    date_order_sheet = sh1["L1"].value
    print("EXCEL ORDER SHEET ", Color.BOLD + date_order_sheet + Color.END, '\n')

    start = input("Start Taking Order? [Y/N]")
    if (start.upper() == "Y") or (start == "1"):
        print("Let's Start taking orders,", name, "!\n")
    else:
        print('Bye,', name, '!')
        exit()

    print(Color.BOLD + "[1] 'A1 KTRA-IND'\n"
                       "[2] 'A2 GANJ'\n"
                       "[3] 'A3 BIDUPUR'\n"
                       "[4] 'A4 MARAI'" + Color.END)
    sh = int(input("Enter 1 or 2 or 3 or 4:\n"))

    row = 5
    column = 1
    count = 1

    if sh == 1:
        sh1_list = []
        while row < 33:
            sh1_list.append([count, sh1.cell(row, 2).value, sh1.cell(row, 3).value])
            row += 1
            count += 1

        count = 0
        while count < 30:
            print(tabulate(sh1_list, headers=["SN No.", "Store", "Phone No."]))
            ans4 = int(input('\nENTER NUMBER: \n')) - 1
            resp = client.messages.create(
                body="\nStore: " + str(sh1_list[ans4][1]) + "\nCall: " + str(sh1_list[ans4][2]),
                from_='+13126754624', to=sendto)
            print('---NOTIFICATION SENT---')

            print('\n\n' + Color.BOLD + sh1_list[ans4][1] + '\n' + str(sh1_list[ans4][2]) + Color.END)
            print('\nEnter Orders: \n')
            count2 = 0
            print(Color.BOLD + '[0] No Order')
            while count2 < len(rate):
                print('[' + str(count2 + 1) + '] ' + str(rate[count2][0]) + ' (' + str(rate[count2][1]) + ') - ' + str(
                    rate[count2][2]) + Color.END)
                count2 += 1

            count3 = 0
            while count3 == 0:
                order = int(input(Color.BOLD + '\nEnter(0-15) :\n' + Color.END))
                row1 = ans4 + 5
                g500 = 0.0
                g1000 = 0.0
                s500 = 0.0
                s1000 = 0.0
                c500 = 0.0
                c1000 = 0.0
                t500 = 0.0
                t1000 = 0.0
                kool_plastic = 0.0
                d200 = 0.0
                d400 = 0.0
                d1000 = 0.0
                lp180 = 0.0
                p200 = 0.0
                p1000 = 0.0

                if order == 0:
                    sh1.cell(row=row1, column=4, value='0')
                if order == 1:
                    g500 = int(input('Enter no. of Gold - 500ml: \n'))
                    sh1.cell(row=row1, column=4, value=g500)
                if order == 2:
                    g1000 = float(input('Enter no. of Gold - 1L: \n'))
                    sh1.cell(row=row1, column=5, value=g1000)
                if order == 3:
                    s500 = float(input('Enter no. of Shakti - 500ml: \n'))
                    sh1.cell(row=row1, column=6, value=s500)
                if order == 4:
                    s1000 = float(input('Enter no. of Shakti - 1L: \n'))
                    sh1.cell(row=row1, column=7, value=s1000)
                if order == 5:
                    c500 = float(input('Enter no. of Cow - 500ml: \n'))
                    sh1.cell(row=row1, column=8, value=c500)
                if order == 6:
                    c1000 = float(input('Enter no. of Cow - 1L: \n'))
                    sh1.cell(row=row1, column=9, value=c1000)
                if order == 7:
                    t500 = float(input('Enter no. of Taaza - 500ml: \n'))
                    sh1.cell(row=row1, column=10, value=t500)
                if order == 8:
                    t1000 = float(input('Enter no. of Taaza - 1L: \n'))
                    sh1.cell(row=row1, column=11, value=t1000)
                if order == 9:
                    kool_plastic = float(input('Enter no. of Amul Kool (Plastic) - 200ml X 30: \n'))
                    sh1.cell(row=row1, column=12, value=kool_plastic)
                if order == 10:
                    d200 = float(input('Enter no. of Dahi - 200ml: \n'))
                    sh1.cell(row=row1, column=13, value=d200)
                if order == 11:
                    d400 = float(input('Enter no. of Dahi - 400ml: \n'))
                    sh1.cell(row=row1, column=14, value=d400)
                if order == 12:
                    d1000 = float(input('Enter no. of Dahi - 1KG: \n'))
                    sh1.cell(row=row1, column=15, value=d1000)
                if order == 13:
                    lp180 = float(input('Enter no. of Lassi(Packet) - 180ml: \n'))
                    sh1.cell(row=row1, column=16, value=lp180)
                if order == 14:
                    p200 = float(input('Enter no. of Paneer - 200g: \n'))
                    sh1.cell(row=row1, column=17, value=p200)
                if order == 15:
                    p1000 = float(input('Enter no. of Paneer - 1KG: \n'))
                    sh1.cell(row=row1, column=18, value=p1000)
                skip = input('More Orders from ' + sh1_list[ans4][1] + '? [Y/N]')
                if (skip.upper() == "Y") or (skip == "1"):
                    count3 = 0
                else:
                    try:
                        wb.save("Excel Files/" + excel_path)
                        print(Color.GREEN+Color.BOLD+'\nSAVED\n'+Color.END)
                    except:
                        print(Color.RED + 'Close excel if open.' + Color.END)
                    count3 = 1

                    if sh1.cell(row1, 4) != 0:
                        print('Calculating Cost...')
                        cost_of_order = g500 * rate[0][2] + g1000 * rate[1][2] + s500 * rate[2][2] + s1000 * rate[3][
                            2] + c500 * rate[4][2] + c1000 * rate[5][2] + t500 * rate[6][2] + t1000 * rate[7][
                                            2] + kool_plastic * rate[8][2] + d200 * rate[9][2] + d400 * rate[10][
                                            2] + d1000 * rate[11][2] + lp180 * rate[12][2] + p200 * rate[13][
                                            2] + p1000 * rate[14][2]
                        print(Color.BLUE+Color.BOLD+cost_of_order+Color.END)
                    break

            count += 1

        else:
            print('Attempts Exceeded, Try Again')

    if sh == 2:
        print('Connected to', sh2)
        sh2_list = []
        while row < 33:
            sh2_list.append([count, sh2.cell(row, 2).value, sh2.cell(row, 3).value])
            row += 1
            count += 1

        count = 0
        while count < 30:
            print(tabulate(sh2_list, headers=["SN No.", "Store", "Phone No."]))
            ans4 = int(input('\nENTER NUMBER: \n')) - 1
            resp = client.messages.create(
                body="\nStore: " + str(sh2_list[ans4][1]) + "\nCall: " + str(sh2_list[ans4][2]),
                from_='+13126754624', to=sendto)
            print('---NOTIFICATION SENT---')

            print('\n\n' + Color.BOLD + sh2_list[ans4][1] + '\n' + str(sh2_list[ans4][2]) + Color.END)
            print('\nEnter Orders: \n')
            count2 = 0
            print(Color.BOLD + '[0] No Order')
            while count2 < len(rate):
                print('[' + str(count2 + 1) + '] ' + str(rate[count2][0]) + ' (' + str(rate[count2][1]) + ') - ' + str(
                    rate[count2][2]) + Color.END)
                count2 += 1

            count3 = 0
            while count3 == 0:
                order = int(input(Color.BOLD + '\nEnter(0-15) :\n' + Color.END))
                row1 = ans4 + 5
                g500 = 0.0
                g1000 = 0.0
                s500 = 0.0
                s1000 = 0.0
                c500 = 0.0
                c1000 = 0.0
                t500 = 0.0
                t1000 = 0.0
                kool_plastic = 0.0
                d200 = 0.0
                d400 = 0.0
                d1000 = 0.0
                lp180 = 0.0
                p200 = 0.0
                p1000 = 0.0

                if order == 0:
                    sh2.cell(row=row1, column=4, value='0')
                if order == 1:
                    g500 = int(input('Enter no. of Gold - 500ml: \n'))
                    sh2.cell(row=row1, column=4, value=g500)
                if order == 2:
                    g1000 = float(input('Enter no. of Gold - 1L: \n'))
                    sh2.cell(row=row1, column=5, value=g1000)
                if order == 3:
                    s500 = float(input('Enter no. of Shakti - 500ml: \n'))
                    sh2.cell(row=row1, column=6, value=s500)
                if order == 4:
                    s1000 = float(input('Enter no. of Shakti - 1L: \n'))
                    sh2.cell(row=row1, column=7, value=s1000)
                if order == 5:
                    c500 = float(input('Enter no. of Cow - 500ml: \n'))
                    sh2.cell(row=row1, column=8, value=c500)
                if order == 6:
                    c1000 = float(input('Enter no. of Cow - 1L: \n'))
                    sh2.cell(row=row1, column=9, value=c1000)
                if order == 7:
                    t500 = float(input('Enter no. of Taaza - 500ml: \n'))
                    sh2.cell(row=row1, column=10, value=t500)
                if order == 8:
                    t1000 = float(input('Enter no. of Taaza - 1L: \n'))
                    sh2.cell(row=row1, column=11, value=t1000)
                if order == 9:
                    kool_plastic = float(input('Enter no. of Amul Kool (Plastic) - 200ml X 30: \n'))
                    sh2.cell(row=row1, column=12, value=kool_plastic)
                if order == 10:
                    d200 = float(input('Enter no. of Dahi - 200ml: \n'))
                    sh2.cell(row=row1, column=13, value=d200)
                if order == 11:
                    d400 = float(input('Enter no. of Dahi - 400ml: \n'))
                    sh2.cell(row=row1, column=14, value=d400)
                if order == 12:
                    d1000 = float(input('Enter no. of Dahi - 1KG: \n'))
                    sh2.cell(row=row1, column=15, value=d1000)
                if order == 13:
                    lp180 = float(input('Enter no. of Lassi(Packet) - 180ml: \n'))
                    sh2.cell(row=row1, column=16, value=lp180)
                if order == 14:
                    p200 = float(input('Enter no. of Paneer - 200g: \n'))
                    sh2.cell(row=row1, column=17, value=p200)
                if order == 15:
                    p1000 = float(input('Enter no. of Paneer - 1KG: \n'))
                    sh2.cell(row=row1, column=18, value=p1000)
                skip = input('More Orders from ' + sh2_list[ans4][1] + '? [Y/N]')
                if (skip.upper() == "Y") or (skip == "1"):
                    count3 = 0
                else:
                    try:
                        wb.save("Excel Files/" + excel_path)
                        print(Color.GREEN+Color.BOLD+'\nSAVED\n'+Color.END)
                    except:
                        print(Color.RED + 'Close excel if open.' + Color.END)
                    count3 = 1

                    if sh2.cell(row1, 4) != 0:
                        print('Calculating Cost...')
                        cost_of_order = g500 * rate[0][2] + g1000 * rate[1][2] + s500 * rate[2][2] + s1000 * rate[3][
                            2] + c500 * rate[4][2] + c1000 * rate[5][2] + t500 * rate[6][2] + t1000 * rate[7][
                                            2] + kool_plastic * rate[8][2] + d200 * rate[9][2] + d400 * rate[10][
                                            2] + d1000 * rate[11][2] + lp180 * rate[12][2] + p200 * rate[13][
                                            2] + p1000 * rate[14][2]
                        print(Color.BLUE+Color.BOLD+cost_of_order+Color.END)
                    break

            count += 1

        else:
            print('Attempts Exceeded, Try Again')
    if sh == 3:
        sh3_list = []
        while row < 33:
            sh3_list.append([count, sh3.cell(row, 2).value, sh3.cell(row, 3).value])
            row += 1
            count += 1

        count = 0
        while count < 30:
            print(tabulate(sh3_list, headers=["SN No.", "Store", "Phone No."]))
            ans4 = int(input('\nENTER NUMBER: \n')) - 1
            resp = client.messages.create(
                body="\nStore: " + str(sh3_list[ans4][1]) + "\nCall: " + str(sh3_list[ans4][2]),
                from_='+13126754624', to=sendto)
            print('---NOTIFICATION SENT---')

            print('\n\n' + Color.BOLD + sh3_list[ans4][1] + '\n' + str(sh3_list[ans4][2]) + Color.END)
            print('\nEnter Orders: \n')
            count2 = 0
            print(Color.BOLD + '[0] No Order')
            while count2 < len(rate):
                print('[' + str(count2 + 1) + '] ' + str(rate[count2][0]) + ' (' + str(rate[count2][1]) + ') - ' + str(
                    rate[count2][2]) + Color.END)
                count2 += 1

            count3 = 0
            while count3 == 0:
                order = int(input(Color.BOLD + '\nEnter(0-15) :\n' + Color.END))
                row1 = ans4 + 5
                g500 = 0.0
                g1000 = 0.0
                s500 = 0.0
                s1000 = 0.0
                c500 = 0.0
                c1000 = 0.0
                t500 = 0.0
                t1000 = 0.0
                kool_plastic = 0.0
                d200 = 0.0
                d400 = 0.0
                d1000 = 0.0
                lp180 = 0.0
                p200 = 0.0
                p1000 = 0.0

                if order == 0:
                    sh3.cell(row=row1, column=4, value='0')
                if order == 1:
                    g500 = int(input('Enter no. of Gold - 500ml: \n'))
                    sh3.cell(row=row1, column=4, value=g500)
                if order == 2:
                    g1000 = float(input('Enter no. of Gold - 1L: \n'))
                    sh3.cell(row=row1, column=5, value=g1000)
                if order == 3:
                    s500 = float(input('Enter no. of Shakti - 500ml: \n'))
                    sh3.cell(row=row1, column=6, value=s500)
                if order == 4:
                    s1000 = float(input('Enter no. of Shakti - 1L: \n'))
                    sh3.cell(row=row1, column=7, value=s1000)
                if order == 5:
                    c500 = float(input('Enter no. of Cow - 500ml: \n'))
                    sh3.cell(row=row1, column=8, value=c500)
                if order == 6:
                    c1000 = float(input('Enter no. of Cow - 1L: \n'))
                    sh3.cell(row=row1, column=9, value=c1000)
                if order == 7:
                    t500 = float(input('Enter no. of Taaza - 500ml: \n'))
                    sh3.cell(row=row1, column=10, value=t500)
                if order == 8:
                    t1000 = float(input('Enter no. of Taaza - 1L: \n'))
                    sh3.cell(row=row1, column=11, value=t1000)
                if order == 9:
                    kool_plastic = float(input('Enter no. of Amul Kool (Plastic) - 200ml X 30: \n'))
                    sh3.cell(row=row1, column=12, value=kool_plastic)
                if order == 10:
                    d200 = float(input('Enter no. of Dahi - 200ml: \n'))
                    sh3.cell(row=row1, column=13, value=d200)
                if order == 11:
                    d400 = float(input('Enter no. of Dahi - 400ml: \n'))
                    sh3.cell(row=row1, column=14, value=d400)
                if order == 12:
                    d1000 = float(input('Enter no. of Dahi - 1KG: \n'))
                    sh3.cell(row=row1, column=15, value=d1000)
                if order == 13:
                    lp180 = float(input('Enter no. of Lassi(Packet) - 180ml: \n'))
                    sh3.cell(row=row1, column=16, value=lp180)
                if order == 14:
                    p200 = float(input('Enter no. of Paneer - 200g: \n'))
                    sh3.cell(row=row1, column=17, value=p200)
                if order == 15:
                    p1000 = float(input('Enter no. of Paneer - 1KG: \n'))
                    sh3.cell(row=row1, column=18, value=p1000)
                skip = input('More Orders from ' + sh3_list[ans4][1] + '? [Y/N]')
                if (skip.upper() == "Y") or (skip == "1"):
                    count3 = 0
                else:
                    try:
                        wb.save("Excel Files/" + excel_path)
                        print(Color.GREEN+Color.BOLD+'\nSAVED\n'+Color.END)
                    except:
                        print(Color.RED + 'Close excel if open.' + Color.END)
                    count3 = 1

                    if sh3.cell(row1, 4) != 0:
                        print('Calculating Cost...')
                        cost_of_order = g500 * rate[0][2] + g1000 * rate[1][2] + s500 * rate[2][2] + s1000 * rate[3][
                            2] + c500 * rate[4][2] + c1000 * rate[5][2] + t500 * rate[6][2] + t1000 * rate[7][
                                            2] + kool_plastic * rate[8][2] + d200 * rate[9][2] + d400 * rate[10][
                                            2] + d1000 * rate[11][2] + lp180 * rate[12][2] + p200 * rate[13][
                                            2] + p1000 * rate[14][2]
                        print(Color.BLUE+Color.BOLD+cost_of_order+Color.END)
                    break

            count += 1

        else:
            print('Attempts Exceeded, Try Again')
    if sh == 4:
        sh4_list = []
        while row < 33:
            sh4_list.append([count, sh4.cell(row, 2).value, sh4.cell(row, 3).value])
            row += 1
            count += 1

        count = 0
        while count < 30:
            print(tabulate(sh4_list, headers=["SN No.", "Store", "Phone No."]))
            ans4 = int(input('\nENTER NUMBER: \n')) - 1
            resp = client.messages.create(
                body="\nStore: " + str(sh4_list[ans4][1]) + "\nCall: " + str(sh4_list[ans4][2]),
                from_='+13126754624', to=sendto)
            print('---NOTIFICATION SENT---')

            print('\n\n' + Color.BOLD + sh4_list[ans4][1] + '\n' + str(sh4_list[ans4][2]) + Color.END)
            print('\nEnter Orders: \n')
            count2 = 0
            print(Color.BOLD + '[0] No Order')
            while count2 < len(rate):
                print('[' + str(count2 + 1) + '] ' + str(rate[count2][0]) + ' (' + str(rate[count2][1]) + ') - ' + str(
                    rate[count2][2]) + Color.END)
                count2 += 1

            count3 = 0
            while count3 == 0:
                order = int(input(Color.BOLD + '\nEnter(0-15) :\n' + Color.END))
                row1 = ans4 + 5
                g500 = 0.0
                g1000 = 0.0
                s500 = 0.0
                s1000 = 0.0
                c500 = 0.0
                c1000 = 0.0
                t500 = 0.0
                t1000 = 0.0
                kool_plastic = 0.0
                d200 = 0.0
                d400 = 0.0
                d1000 = 0.0
                lp180 = 0.0
                p200 = 0.0
                p1000 = 0.0

                if order == 0:
                    sh4.cell(row=row1, column=4, value='0')
                if order == 1:
                    g500 = int(input('Enter no. of Gold - 500ml: \n'))
                    sh4.cell(row=row1, column=4, value=g500)
                if order == 2:
                    g1000 = float(input('Enter no. of Gold - 1L: \n'))
                    sh4.cell(row=row1, column=5, value=g1000)
                if order == 3:
                    s500 = float(input('Enter no. of Shakti - 500ml: \n'))
                    sh4.cell(row=row1, column=6, value=s500)
                if order == 4:
                    s1000 = float(input('Enter no. of Shakti - 1L: \n'))
                    sh4.cell(row=row1, column=7, value=s1000)
                if order == 5:
                    c500 = float(input('Enter no. of Cow - 500ml: \n'))
                    sh4.cell(row=row1, column=8, value=c500)
                if order == 6:
                    c1000 = float(input('Enter no. of Cow - 1L: \n'))
                    sh4.cell(row=row1, column=9, value=c1000)
                if order == 7:
                    t500 = float(input('Enter no. of Taaza - 500ml: \n'))
                    sh4.cell(row=row1, column=10, value=t500)
                if order == 8:
                    t1000 = float(input('Enter no. of Taaza - 1L: \n'))
                    sh4.cell(row=row1, column=11, value=t1000)
                if order == 9:
                    kool_plastic = float(input('Enter no. of Amul Kool (Plastic) - 200ml X 30: \n'))
                    sh4.cell(row=row1, column=12, value=kool_plastic)
                if order == 10:
                    d200 = float(input('Enter no. of Dahi - 200ml: \n'))
                    sh4.cell(row=row1, column=13, value=d200)
                if order == 11:
                    d400 = float(input('Enter no. of Dahi - 400ml: \n'))
                    sh4.cell(row=row1, column=14, value=d400)
                if order == 12:
                    d1000 = float(input('Enter no. of Dahi - 1KG: \n'))
                    sh4.cell(row=row1, column=15, value=d1000)
                if order == 13:
                    lp180 = float(input('Enter no. of Lassi(Packet) - 180ml: \n'))
                    sh4.cell(row=row1, column=16, value=lp180)
                if order == 14:
                    p200 = float(input('Enter no. of Paneer - 200g: \n'))
                    sh4.cell(row=row1, column=17, value=p200)
                if order == 15:
                    p1000 = float(input('Enter no. of Paneer - 1KG: \n'))
                    sh4.cell(row=row1, column=18, value=p1000)
                skip = input('More Orders from ' + sh4_list[ans4][1] + '? [Y/N]')
                if (skip.upper() == "Y") or (skip == "1"):
                    count3 = 0
                else:
                    try:
                        wb.save("Excel Files/" + excel_path)
                        print(Color.GREEN+Color.BOLD+'\nSAVED\n'+Color.END)
                    except:
                        print(Color.RED + 'Close excel if open.' + Color.END)
                    count3 = 1

                    if sh4.cell(row1, 4) != 0:
                        print('Calculating Cost...')
                        cost_of_order = g500 * rate[0][2] + g1000 * rate[1][2] + s500 * rate[2][2] + s1000 * rate[3][
                            2] + c500 * rate[4][2] + c1000 * rate[5][2] + t500 * rate[6][2] + t1000 * rate[7][
                                            2] + kool_plastic * rate[8][2] + d200 * rate[9][2] + d400 * rate[10][
                                            2] + d1000 * rate[11][2] + lp180 * rate[12][2] + p200 * rate[13][
                                            2] + p1000 * rate[14][2]
                        print(Color.BLUE+Color.BOLD+cost_of_order+Color.END)
                    break

            count += 1

        else:
            print('Attempts Exceeded, Try Again')

else:
    print('Great, ' + name)
    exit()
