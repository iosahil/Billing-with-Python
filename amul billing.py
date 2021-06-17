# Importing Libraries
from tabulate import tabulate
import datetime as datetime
from tqdm import tqdm as progressbar  # Visual using progress bar
import time
import pyfiglet as art
import openpyxl as excel


class Color:
    PURPLE = '\033[95m'
    CYAN = '\033[96m'
    DARKCYAN = '\033[36m'
    BLUE = '\033[94m'
    GREEN = '\033[92m'
    YELLOW = '\033[93m'
    RED = '\033[91m'
    BOLD = '\033[1m'
    UNDERLINE = '\033[4m'
    END = '\033[0m'


name = str(input("\nEnter your name:"))
name = name.capitalize()
print('\nHi,', Color.BOLD + name.capitalize() + Color.END)

passcode = 0x75AB
count = 0
user_pass = ""
while user_pass != passcode and count < 3:
    try:
        user_pass = int(input('Enter PIN:\n'))
    except:
        print("\nERROR\n•Password can't consists of alphabets.\n•Password must be 5 digits.\n")

    if user_pass == passcode:
        print('Access granted for', name.capitalize() + ".\n")
        break
    else:
        print('Access denied. Try again ' + name.capitalize() + ".")
        count += 1
        print('You have', 3 - count, 'counts left.\n')
        if count == 3:
            print('Tumse na ho payega', name.lower(), ':)')
            print('\033[1m' + 'USER LOCKED OUT!!')
            exit()

print(Color.BOLD + 'WELCOME TO BILLING MACHINE, ' + name.upper() + Color.END)

rate = [['Gold', '500ml', 26.18],
        ['Gold', '1L', 51.35],
        ['Shakti', '500ml', 23.20],
        ['Shakti', '1L', 45.40],
        ['Cow', '500ml', 22.20],
        ['Cow', '1L', 43.40],
        ['Taaza', '500ml', 21.18],
        ['Taaza', '1L', 41.40],
        ['Amul Kool (Plastic)', '200ml * 30', 540.00],
        ['Dahi', '200g', 13.86],
        ['Dahi', '400g', 24.63],
        ['Dahi', '1KG', 56.80],
        ['Lassi (Packet)', '180ml', 9.10],
        ['Paneer', '200g', 66.00],
        ['Paneer', '1KG', 315.00]
        ]

ans1 = input('Want to know rates? [Y/N] \n')
if (ans1 == "Y") or (ans1 == "y"):
    print(tabulate(rate, headers=[Color.BOLD + "Item", "Quantity", "Selling Price" + Color.END], numalign="left"))
    print('\nBahut sasta hai, \nWholesale price hai!!!! \n#Exclusively_On_Softline')
else:
    print('Okay,', name, ':)')

ans2 = input('\nWant to start billing machine? [Y/N] \n')
if (ans2 == "Y") or (ans2 == "y"):
    try:
        for i in progressbar(range(10),
                             desc='Opening with superfast speed',
                             ascii=False, ncols=75):
            time.sleep(0.25)
        print("Complete.\n")
    except:
        print(Color.RED + "\nERROR ~~ 'tqdm' package not found\nSkipping progress bar visualization\n\n" + Color.END)

    # Let's create art
    try:
        banner = art.figlet_format("WELCOME TO SOFTLINE AMUL BILLING", font="digital")
        print(Color.BOLD + Color.GREEN + banner + Color.END)
    except:
        print(Color.RED + "\nERROR ~~ 'pyfiglet' package not found\nSkipping art visualization\n\n" + Color.END)
        print('Opening Billing Machine...')

    # Excel Connection Starts Here!
    print('Connect to EXCEL of date:\n')
    print(Color.BOLD + "[1] Today\n"
                       "[2] Yesterday\n"
                       "[3] Enter Date Manually\n" + Color.END)
    ans3 = int(input("Enter 1 or 2 or 3:\n"))
    dat = ''
    if ans3 == 1:
        dat = today = datetime.date.today().strftime("%d-%m-%Y")
    if ans3 == 2:
        yesterday = datetime.date.today() - datetime.timedelta(days=1)
        dat = yesterday.strftime("%d-%m-%Y")
    if ans3 == 3:
        dat = manual_date = str(input("\nEnter Date Manually (DD-MM-YYYY):"))
    try:
        excel_path = 'O-' + dat + '.xlsx'
        wb = excel.load_workbook("Excel Files\\" + excel_path)
    except:
        print(Color.RED + "ERROR\n"
                          "Can't find Excel sheet with date - " + dat + ".\n"
                                                                        "Please Enter Excel File Location Manually." + Color.END)
        excel_path = input("\n\nEnter Excel file path \n(without quotes)(with double slashes):\n")
        wb = excel.load_workbook(excel_path)
    try:
        print('Connected with Excel - ' +Color.BOLD + excel_path+Color.END)
        print('Connecting with sheets inside Excel...')
        sh1 = wb["A1 KTRA-IND"]
        sh2 = wb["A2 GANJ"]
        sh3 = wb["A3 BIDUPUR"]
        sh4 = wb['A4 MARAI']
        print(Color.BLUE + Color.BOLD + 'File successfully connected.\n' + Color.END)
    except:
        print(Color.RED + "ERROR: File path is invalid!" + Color.END)
        exit()

    date_order_sheet = sh1["L1"].value
    print("Excel Order Sheet Date is ", date_order_sheet, '\n')

    start = input("Start Taking Order? [Y/N]")
    if (start == "Y") or (start == "y"):
        print("Let's Start taking orders,", name, "!")
    else:
        print('Bye,', name, '!')
        exit()

    print(Color.BOLD+"[1]+ 'A1 KTRA-IND'\n"
          "[2] 'A2 GANJ'\n"
          "[3] 'A3 BIDUPUR'\n"
          "[4] 'A4 MARAI'"+Color.END)
    sh = int(input("Enter 1 or 2 or 3 or 4:\n"))

    row = 5
    column = 1
    count = 1

    sh1_list = []
    while row < 33:
        sh1_list.append([count, sh1.cell(row, 2).value, sh1.cell(row, 3).value], )
        try:
            unnecessary = sh1_list.index([count, None, None])
            print(unnecessary)
            # count -= 1
        except:
            ''

        row += 1
        count += 1

    print(sh1_list)
    if sh == 1:
        print('Connected to', sh1)
        while row < 33:
            print(count)
            print("Store: ", sh1.cell(row, 2).value)
            print("Call: ", sh1.cell(row, 3).value, "\n")

            count += 1
            row += 1
    if sh == 2:
        print('Connected to', sh2)
        while row < 33:
            print(count)
            print("Store: ", sh2.cell(row, 2).value)
            print("Call: ", sh2.cell(row, 3).value, "\n")
            # print(tabulate(sh2_list, headers=["Item", "Quantity", "Selling Price"], numalign="left"))
            count += 1
            row += 1
    if sh == 3:
        print('Connected to', sh3)
        while row < 33:
            print(count)
            print("Store: ", sh2.cell(row, 2).value)
            print("Call: ", sh2.cell(row, 3).value, "\n")
            count += 1
            row += 1
    if sh == 4:
        print('Connected to', sh4)
        while row < 33:
            print(count)
            print("Store: ", sh4.cell(row, 2).value)
            print("Call: ", sh4.cell(row, 3).value, "\n")
            count += 1
            row += 1

else:
    print('Gaand mara bsdk ' + name)
    exit()
